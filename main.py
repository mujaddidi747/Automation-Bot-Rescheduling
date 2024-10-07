import time
import openpyxl
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException

# Set up Chrome options with the custom user agent
chrome_options = Options()
user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
chrome_options.add_argument(f"user-agent={user_agent}")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option('useAutomationExtension', False)

driver = webdriver.Chrome(options=chrome_options)
new_bookings = []
booking_details = []
bookings_to_schedule = []
delay = 0
option_keywords = ""
testing = ""

def find_smallest_date(dates):
    try:
        date_objects = [datetime.strptime(date, "%Y-%m-%d") for date in dates]
        if date_objects:
            smallest_date = min(date_objects)
            return smallest_date.strftime("%Y-%m-%d")
        else:
            return None
    except ValueError as e:
        print(f"Error parsing dates: {e}")
        return None

def check_element_exists_by_class(class_name):
    try:
        driver.find_element(By.CLASS_NAME, class_name)
        return True
    except NoSuchElementException:
        return False

def get_current_month():
    return datetime.now().strftime("%B")

def clickOnTimeSlot():
    buttons = driver.find_elements(By.CSS_SELECTOR, "div.time-selector button")
    # Filter out the buttons that are disabled
    enabled_buttons = [button for button in buttons if not button.get_attribute("disabled")]
    # Click the first enabled button if any exist
    if enabled_buttons:
        enabled_buttons[0].click()
        print(f"Clicked the first enabled button: {enabled_buttons[0].text}")
    else:
        print("No enabled buttons found.")

def find_and_click_back_button(times):
    status = True
    while status:
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.cal-selector div.filter button.back')))
            back_button = driver.find_element(By.CSS_SELECTOR, 'div.cal-selector div.filter button.back')
            for _ in range(times):
                back_button.click()
                time.sleep(1)
            status = False
        except (NoSuchElementException, StaleElementReferenceException) as e:
            print(f"Failed to click the 'back' button: {e}")
            time.sleep(5)
        
def find_and_click_next_button(times):
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.cal-selector div.filter button.next')))
        next_button = driver.find_element(By.CSS_SELECTOR, 'div.cal-selector div.filter button.next')
        for _ in range(times):
            next_button.click()
            time.sleep(1)
    except (NoSuchElementException, StaleElementReferenceException) as e:
        print(f"Failed to click the 'next' button: {e}")

def get_item_texts():
    available_days = []
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.body-ct div.item')))
        items = driver.find_elements(By.CSS_SELECTOR, 'div.body-ct div.item')
        for item in items:
            try:
                wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'available')))
                available_spans = item.find_elements(By.CLASS_NAME, 'available')
                for span in available_spans:
                    if 'disabled' not in span.get_attribute('class'):
                        available_days.append(span.text)
            except (NoSuchElementException, StaleElementReferenceException) as e:
                print(f"Error processing items: {e}")
    except NoSuchElementException:
        print("No items found with the class 'item' inside 'body-ct'")
    return available_days


def clickOnDate(date):
    try:
        date_object = datetime.strptime(date, "%Y-%m-%d")
        day = date_object.day

        items = driver.find_elements(By.CSS_SELECTOR, 'div.body-ct div.item')
        for item in items:
            try:
                available_spans = item.find_elements(By.CLASS_NAME, 'available')
                for span in available_spans:
                    if 'disabled' not in span.get_attribute('class'):
                        if int(span.text.strip()) == int(day):
                            span.click()
            except (NoSuchElementException, StaleElementReferenceException) as e:
                print(f"Error processing items: {e}")
    except NoSuchElementException:
        print("No items found with the class 'item' inside 'body-ct'")

def addVehicleDetails(booking_data):
    time.sleep(4)
    clicked = True
    while clicked != False:
        try:
            labels = driver.find_elements(By.CSS_SELECTOR, "div.vehicles-tabs div.vehicle-types label")
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.vehicles-tabs div.vehicle-types label")))

            if labels:
                labels[0].click()
                clicked = False
                print("Clicked on the first label:", labels[0].text)
            else:
                print("No labels found.")
        except NoSuchElementException:
            continue
        time.sleep(3)
        
        
    addInput(str(booking_data[9]), "div.vehicles-tabs input[title='VIN/Chassis Number']")
    addInput(str(booking_data[10]), "div.vehicles-tabs input[title='Vehicle Manufacturer']")
    addInput(str(booking_data[11]), "div.vehicles-tabs input[title='Vehicle Model']")
    addInput(str(booking_data[13]), "div.vehicles-tabs input[title='Vehicle Colour']")
    addInput(str(booking_data[16]), "div.vehicles-tabs input[title='Where was this vehicle purchased']")
    
    selectOption(str(booking_data[12]), "div.vehicles-tabs select[title='Year']")
    selectOption("Jan", "div.vehicles-tabs select[title='Month']")
    selectOption(str(booking_data[14]), "div.vehicles-tabs select[title='WOVR Damage Description']")
    selectOption(str(booking_data[15]), "div.vehicles-tabs select[title=\"Vehicle's Purchase Method\"]")

    clickButton("form[name='formVehicle'] div.form-action button")

def addCustomerDetails(booking_data):
    index = getCustomerDataIndex(booking_data[5])
    
    addInput(str(booking_details[0][index]), "form[id='form-customer'] input[title='Queensland CRN']")
    addInput(str(booking_details[1][index]), "form[id='form-customer'] input[title='First Name']")
    addInput(str(booking_details[2][index]), "form[id='form-customer'] input[title='Last Name']")
    addInput(str(booking_details[3][index]), "form[id='form-customer'] input[title='Street Address']")
    addInput(str(booking_details[4][index]), "form[id='form-customer'] input[title='Suburb']")
    addInput(str(booking_details[5][index]), "form[id='form-customer'] input[title='Post Code']")
    addInput(str(booking_details[6][index]), "form[id='form-customer'] input[title='Email Address']")
    addInput(str(booking_details[7][index]), "form[id='form-customer'] input[title='Phone Number']")
    
    clickButton("form[id='form-customer'] div.form-action button")

def bookingConfirm():
    clickButton("form[id='form-confirm'] button[id='Paperwork']")
    
    time.sleep(2)
    try:
        h4_elements = driver.find_elements(By.CSS_SELECTOR, "form#form-confirm h4")
        for h4 in h4_elements:
            if h4.text.strip() == "Booking Fee":
                driver.execute_script("arguments[0].scrollIntoView();", h4)
                print("Found h4 text:", h4.text)
                break
    except NoSuchElementException:
        print(f"The element with css_selector '{css_selector}' was not found.")    
    time.sleep(1)
    
    iframe = driver.find_element(By.XPATH, "//iframe[@title='reCAPTCHA']")
    driver.switch_to.frame(iframe)
    rc_anchor_container = driver.find_element(By.ID, "rc-anchor-container")
    rc_anchor_container.click()

    driver.switch_to.default_content()    

def clickSubmit():
    clickButton("form[id='form-confirm'] div[class='form-action'] button[type='submit']")

def clickButton(css_selector):
    button = driver.find_element(By.CSS_SELECTOR, css_selector)
    if button:
        try:
            driver.execute_script("arguments[0].scrollIntoView();", button)
        except NoSuchElementException:
            print(f"The element with css_selector '{css_selector}' was not found.")
        button.click()

def addInput(text, css_selector):
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, css_selector)))
    vehicle_purchased = driver.find_element(By.CSS_SELECTOR, css_selector)
    vehicle_purchased.clear()
    vehicle_purchased.send_keys(text)
    
def selectOption(option_name, css_selector):
    try:
        select_element = driver.find_element(By.CSS_SELECTOR, css_selector)
        select = Select(select_element)

        index_of_label = -1
        for index, option in enumerate(select.options):
            if option.text.lower() == option_name.lower():
                index_of_label = index
                break

        if index_of_label != -1:
            select.select_by_index(index_of_label)
    except NoSuchElementException:
        print(f"The select with css_selector '{css_selector}' was not found.")

def bookAppointment(earliest_date_key, earliest_date_value):
    try:
        select_element = driver.find_element(By.CSS_SELECTOR, 'select[title="Select a Location to view Booking Availability"]')
        select = Select(select_element)

        label_to_find = earliest_date_key
        index_of_label = -1
        for index, option in enumerate(select.options):
            if option.text == label_to_find:
                index_of_label = index
                break

        if index_of_label != -1:
            try:
                select.select_by_index(index_of_label)
                time.sleep(1)

                if check_element_exists_by_class("body-ct"):
                    try:
                        element = driver.find_element(By.CLASS_NAME, 'body-ct')
                        driver.execute_script("arguments[0].scrollIntoView();", element)
                    except NoSuchElementException:
                        print("The element with ID 'body-ct' was not found.")

                time.sleep(1)

                filter_text_span = driver.find_element(By.CSS_SELECTOR, 'div.cal-selector div.filter span.filter-text')
                filter_text = filter_text_span.text
                
                # Parse the date strings to datetime objects
                date1 = datetime.strptime(filter_text, "%B %Y")
                date2 = datetime.strptime(earliest_date_value, "%Y-%m-%d")
                # Calculate the difference in months
                year_diff = date2.year - date1.year
                month_diff = date2.month - date1.month
                # Total difference in months
                total_month_diff = year_diff * 12 + month_diff

                if total_month_diff < 0:
                    find_and_click_back_button(abs(total_month_diff))
                    print("click back button")
                elif total_month_diff > 0:
                    find_and_click_next_button(abs(total_month_diff))
                    print("click next button")
                
                clickOnDate(earliest_date_value)

                time.sleep(1)

                clickOnTimeSlot()

                time.sleep(1)

                button = driver.find_element(By.CSS_SELECTOR, "#form-time-loc div.form-action button")
                try:
                    driver.execute_script("arguments[0].scrollIntoView();", button)
                except NoSuchElementException:
                    print("The element with ID 'body-ct' was not found.")
                button.click()
                
                print("Complted!")
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred while selecting an option.")
        
    except NoSuchElementException:
        print("The select element for locations was not found.")
        # driver.quit()
        # exit()

def is_valid_datetime(date_obj):
    if isinstance(date_obj, datetime):
        return True  # It's already a valid datetime object
    return False

def checkIfEligible(date):
    current_date = datetime.now().strftime("%Y-%m-%d")
    booking_date = date.strftime("%Y-%m-%d")

    if booking_date > current_date:
        return True
    else:
        return False
    
def loadNewBookingData():
    workbook = load_workbook('NewBookings.xlsx')
    sheet = workbook.worksheets[0]
    
    for row in sheet.iter_rows(values_only=True):
        new_bookings.append(list(row))

def checkIfNotBooked(vin_number):
    for row in new_bookings:
        if row[9] == vin_number:
            return False
    return True

def loadBookingData():
    workbook = load_workbook('Bookings.xlsx')
    sheet = workbook.worksheets[0]
    for row in sheet.iter_rows(values_only=True):
        booking_details.append(list(row))
    
    sheet = workbook.worksheets[1]
    i = 0
    for row in sheet.iter_rows(values_only=True):
        if i > 0 and is_valid_datetime(list(row)[1]):
            if checkIfEligible(list(row)[1]) and checkIfNotBooked(list(row)[9]):
                bookings_to_schedule.append(list(row))
                # print(list(row))
        i = i + 1

def getBookingData():
    loadNewBookingData()
    loadBookingData()
    if bookings_to_schedule:
        return bookings_to_schedule[0]
    else:
        return False

def updateBookingsFile(data):
    workbook = openpyxl.load_workbook("NewBookings.xlsx")
    sheet = workbook.active

    sheet.append(data)

    workbook.save("NewBookings.xlsx")
    print(f"Data appended to {"NewBookings.xlsx"}")


def getCustomerDataIndex(email):
    index = -1
    i = 0
    for data in booking_details[5]:
        if str(data).strip() == str(email).strip():
            index = i
        i = i + 1
        
    return index
    
    
def getOptionIndex(value, select):
    for index, option in enumerate(select.options):
        if option.get_attribute("value") == value:
            print(f"Index: {index}")
            return 3
        
    return False
            

def submit():
    global delay, option_keywords
    
    delay = int(delay_entry.get())
    
    option_keywords = [keyword.strip() for keyword in keyword_text.get("1.0", "end").strip().split("\n") if keyword.strip()]
        
    if len(option_keywords) == 0:
        messagebox.showerror("Input Error", "Please enter at least one keyword!")
        return
    
    root.destroy()
    
    


# ------------------------Functions Above---------------------------
    
root = tk.Tk()
root.title("Delay and Keywords Input")
root.geometry("400x400")

# Styling
root.configure(bg="#f2f2f2")
title_label = tk.Label(root, text="Enter Delay and Keywords", font=("Helvetica", 16), bg="#f2f2f2")
title_label.pack(pady=10)

# Delay input field
delay_label = tk.Label(root, text="Delay (seconds):", font=("Helvetica", 12), bg="#f2f2f2")
delay_label.pack(pady=5)
delay_entry = tk.Entry(root, width=10, textvariable=tk.StringVar(value="10"), font=("Helvetica", 12))
delay_entry.pack(pady=5)

# Keywords text area
keyword_label = tk.Label(root, text="Keywords (each on a new line):", font=("Helvetica", 12), bg="#f2f2f2")
keyword_label.pack(pady=5)
keyword_text = tk.Text(root, height=8, width=40, font=("Helvetica", 12))
keyword_text.pack(pady=5)

testing = tk.IntVar(value=1)
checkbox = tk.Checkbutton(root, text="Testing Mode", variable=testing)
checkbox.pack()
checkbox.pack(pady=5)
checkbox.pack(padx=20)

# Submit button
submit_button = tk.Button(root, text="Submit", font=("Helvetica", 12), command=submit, bg="#4CAF50", fg="white")
submit_button.pack(pady=10)

# Start the GUI event loop
root.mainloop()

time.sleep(2)


global wait

try:
    booking_data = getBookingData()

    if booking_data == False:
        print("No New Appointments to Reschedule!")
            
    while booking_data != False:
        wait = WebDriverWait(driver, 20)
        
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"Current Time: {current_time}")
        
        
        # Open the website
        driver.get('https://www.wovi.com.au/bookings/')  # Replace with the target website URL

        # Scroll to the element with ID 'form-time-loc'
        try:
            wait.until(EC.presence_of_element_located((By.ID, "form-time-loc")))
            element = driver.find_element(By.ID, 'form-time-loc')
            driver.execute_script("arguments[0].scrollIntoView();", element)
        except NoSuchElementException:
            print("The element with ID 'form-time-loc' was not found.")

        # Find the select element by its title
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'select[title="Select a Location to view Booking Availability"]')))
            select_element = driver.find_element(By.CSS_SELECTOR, 'select[title="Select a Location to view Booking Availability"]')
            select = Select(select_element)
        except NoSuchElementException:
            print("The select element for locations was not found.")
            driver.quit()
            exit()

        options = select.options
        all_dates = {}
        
        
        print(f"option_keywords::: {option_keywords}")
        # Iterate over each option and select it
        for option_keyword in option_keywords:
            try:
                select.select_by_visible_text(option_keyword)
                print(f"Keywords:::: {option_keyword}")
                time.sleep(2)
                selected_option = select.first_selected_option
                print(f"\n--------------------------\nLocation: {selected_option.text}\n")
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred while selecting an option.")
                continue

            if check_element_exists_by_class("body-ct"):
                try:
                    element = driver.find_element(By.CLASS_NAME, 'body-ct')
                    driver.execute_script("arguments[0].scrollIntoView();", element)
                except NoSuchElementException:
                    print("The element with ID 'body-ct' was not found.")
                
                
                dates = []
                month_found = False
                while not month_found:
                    try:
                        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.cal-selector div.filter span.filter-text')))
                        filter_text_span = driver.find_element(By.CSS_SELECTOR, 'div.cal-selector div.filter span.filter-text')
                        filter_text = filter_text_span.text
                        
                        date_object = datetime.strptime(filter_text, "%B %Y")
                        formatted_date = date_object.strftime("%Y-%m")
                        available_days = get_item_texts()

                        if available_days:
                            closest_day = min(available_days)
                            dates.append(formatted_date + '-' + closest_day)
                        
                        current_month = get_current_month()
                        if current_month in filter_text:
                            month_found = True
                        else:
                            find_and_click_back_button(1)
                    except (NoSuchElementException, StaleElementReferenceException) as e:
                        print(f"Error in processing calendar: {e}")
                        break
                
                closest_date = find_smallest_date(dates)
                if closest_date:
                    print(f"Closest Date: {closest_date}")
                    all_dates[selected_option.text] = closest_date
            else:
                print("This location does not have an appointment booking calendar.")

        if all_dates:
            earliest_date_key = min(all_dates, key=all_dates.get)
            earliest_date_value = all_dates[earliest_date_key]
            booking_data.append(earliest_date_value)
            booking_data[3] = earliest_date_key
            print(f"-------\nLocation: {earliest_date_key}")
            print(f"Date: {earliest_date_value}")
            
            date_format = "%Y-%m-%d"
            closest_date_obj = datetime.strptime(earliest_date_value, date_format)
            print(closest_date_obj)
            print(booking_data[1])
            if closest_date_obj < booking_data[1]:
                print("Yes")
                time.sleep(1)
                
                bookAppointment(earliest_date_key, earliest_date_value)
                addVehicleDetails(booking_data)
                time.sleep(1)
                addCustomerDetails(booking_data)
                time.sleep(1)
                bookingConfirm()
                if testing.get() == 0:
                    clickSubmit()
                    updateBookingsFile(booking_data)
            else:
                print("No")
        else:
            print("No dates were found.")


        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"Current Time: {current_time}")
        time.sleep(delay)
        
        new_bookings = []
        booking_details = []
        bookings_to_schedule = []
        booking_data = []
        booking_data = getBookingData()

    # time.sleep(500)
finally:
    # Close the browser
    driver.quit()